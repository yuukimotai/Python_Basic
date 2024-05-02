import openpyxl as px

#wb = px.Workbook()
wb = px.load_workbook('excel_sample.xlsx')
# ws = wb.active
# ws['A1'] = 'Hello Excel!'# ExcelのA1セルに文字列
# ws.title = 'sample_title1'# Excelファイル内のシートのタイトル
ws = wb.create_sheet()
ws.title = 'sample_chart'

ws.append(['Data1'])

for i in range(10):
  ws.append([i + 2])
values = px.chart.Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = px.chart.BarChart()
chart.add_data(values, titles_from_data=True)
ws.add_chart(chart, 'E15')
wb.save('excel_sample.xlsx')